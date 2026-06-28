"""
Local component database — parse the Excel parametric tables, map a stored part to an engine
block, filter by the designer's selection criteria, and rank candidates by computed loss.

The Excel DBs (specs/Database) are Digi-Key-style parametric tables: they carry the selection
parameters (voltage/current rating, Rds(on), Qg, Vf, technology, package, mounting, Tj) but not
the full loss-model detail (Eoss curve, Rth_jc, Qc/Qrr, Vf curve). We map the real scalars
verbatim and ESTIMATE the missing loss/thermal parameters with consistent, clearly-labelled
defaults — so the relative ranking is driven by the real datasheet values, and the designer
refines the picked part (or uploads a datasheet) before signing off.
"""
from __future__ import annotations
import os, re, json

_HERE = os.path.dirname(os.path.abspath(__file__))
_DATA = os.path.join(_HERE, "data")
_SPEC = os.path.normpath(os.path.join(_HERE, "..", "..", "..", "..", "specs", "Database"))

_SRC = {
    "bridge": "bridge_rectifiers_combined_sorted.xlsx",
    "mosfet": "mosfets_combined_sorted.xlsx",
    "diode":  "diodes_combined_sorted.xlsx",
}
_PREFIX = {"p": 1e-12, "n": 1e-9, "u": 1e-6, "µ": 1e-6, "μ": 1e-6, "m": 1e-3,
           "": 1.0, "k": 1e3, "K": 1e3, "M": 1e6}

# ── value parsers (return None on failure) ────────────────────────────────────
def _si(s, unit, allow_no_prefix=True):
    if s is None: return None
    pre = "pnuµμmkKM" + ("" if allow_no_prefix else "")
    m = re.search(r"(-?\d+\.?\d*)\s*([" + pre + r"]?)\s*" + re.escape(unit), str(s))
    if not m: return None
    if not allow_no_prefix and m.group(2) == "": return None
    return float(m.group(1)) * _PREFIX.get(m.group(2), 1.0)

def p_volt(s):  return _si(s, "V")
def p_amp(s):   return _si(s, "A")
def p_res(s):
    if s is None: return None
    ohm = "(?:[" + chr(0x03A9) + chr(0x2126) + "]|[oO]hms?)"
    mu = chr(0xb5) + chr(0x3bc)
    m = re.search(r"(-?\d+\.?\d*)\s*([pnu" + mu + r"mkKM]?)\s*" + ohm, str(s))
    return float(m.group(1)) * _PREFIX.get(m.group(2), 1.0) if m else None
def p_time(s):  return _si(s, "s")
def p_power(s): return _si(s, "W")
def p_charge(s):                                   # "17 nC" — require a prefix so "150°C" isn't a hit
    return _si(s, "C", allow_no_prefix=False)
def p_tjmax(s):                                    # max temperature among "NNN°C"
    if s is None: return None
    vals = [float(x) for x in re.findall(r"(-?\d+\.?\d*)\s*°?C", str(s))]
    return max(vals) if vals else None
def p_at(s, left, right):                          # "1.65 V @ 5 A" → (left(=Vf), right(=If))
    if s is None: return None, None
    parts = str(s).split("@")
    a = left(parts[0]) if parts else None
    b = right(parts[1]) if len(parts) > 1 else None
    return a, b
def p_vf(s):    return p_at(s, p_volt, p_amp)       # (Vf, If)
def p_cap(s):   return p_at(s, lambda x: _si(x, "F"), p_volt)   # (Ciss, Vds)


# ── Excel ingest → normalized records ─────────────────────────────────────────
def _rows(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    for r in it:
        if any(v not in (None, "") for v in r):
            yield dict(zip(hdr, r))
    wb.close()

def _common(r):
    return {"mfr": r.get("Mfr"), "part_number": r.get("Mfr Part #"),
            "description": r.get("Description"), "technology": r.get("Technology"),
            "mounting": r.get("Mounting Type"), "package": r.get("Package / Case"),
            "datasheet_url": r.get("Datasheet") or r.get("URL")}

def ingest(kind):
    path = os.path.join(_SPEC, _SRC[kind]); out = []
    for r in _rows(path):
        c = _common(r)
        if not c["part_number"]:
            continue
        if kind == "mosfet":
            ciss, ciss_v = p_cap(r.get("Input Capacitance (Ciss) (Max) @ Vds"))
            vth, _ = p_at(r.get("Vgs(th) (Max) @ Id"), p_volt, p_amp)
            c.update({"vdss": p_volt(r.get("Drain to Source Voltage (Vdss)")),
                      "id_25": p_amp(r.get("Current - Continuous Drain (Id) @ 25°C")),
                      "rdson": p_res(r.get("Rds On (Max) @ Id, Vgs")),
                      "qg": p_charge(r.get("Gate Charge (Qg) (Max) @ Vgs")),
                      "ciss": ciss, "vth": vth,
                      "pd_max": p_power(r.get("Power Dissipation (Max)")),
                      "tj_max": p_tjmax(r.get("Operating Temperature")),
                      "supplier_package": r.get("Supplier Device Package")})
        elif kind == "diode":
            vf, vf_if = p_vf(r.get("Voltage - Forward (Vf) (Max) @ If"))
            c.update({"vr": p_volt(r.get("Voltage - DC Reverse (Vr) (Max)")),
                      "io": p_amp(r.get("Current - Average Rectified (Io)")),
                      "vf": vf, "vf_if": vf_if,
                      "trr": p_time(r.get("Reverse Recovery Time (trr)")),
                      "tj_max": p_tjmax(r.get("Operating Temperature - Junction"))})
        else:  # bridge
            vf, vf_if = p_vf(r.get("Voltage - Forward (Vf) (Max) @ If"))
            c.update({"vr": p_volt(r.get("Voltage - Peak Reverse (Max)")),
                      "io": p_amp(r.get("Current - Average Rectified (Io)")),
                      "vf": vf, "vf_if": vf_if,
                      "tj_max": p_tjmax(r.get("Operating Temperature"))})
        out.append(c)
    return out

def build_all():
    os.makedirs(_DATA, exist_ok=True)
    counts = {}
    for kind in _SRC:
        recs = ingest(kind)
        with open(os.path.join(_DATA, kind + ".json"), "w", encoding="utf-8") as f:
            json.dump(recs, f)
        counts[kind] = len(recs)
    return counts


# ── load + filter ─────────────────────────────────────────────────────────────
_CACHE = {}
def load(kind):
    if kind not in _CACHE:
        with open(os.path.join(_DATA, kind + ".json"), encoding="utf-8") as f:
            _CACHE[kind] = json.load(f)
    return _CACHE[kind]

def _is_sic(rec):
    t = (rec.get("technology") or "")
    return "SiC" in t or "Silicon Carbide" in t

def options(kind):
    """Distinct selectable values (for the GUI dropdowns)."""
    recs = load(kind)
    uniq = lambda key: sorted({(r.get(key) or "").strip() for r in recs if r.get(key)})
    out = {"manufacturers": uniq("mfr"), "mounting": uniq("mounting"), "package": uniq("package")}
    if kind == "mosfet":
        out["technology"] = uniq("technology")
    return out

def filter_parts(kind, crit):
    """crit: {v_min, i_min, mfr, mounting, package, tj_min, technology}."""
    recs = load(kind)
    vkey = "vdss" if kind == "mosfet" else "vr"
    ikey = "id_25" if kind == "mosfet" else "io"
    out = []
    for r in recs:
        v = r.get(vkey); i = r.get(ikey)
        if crit.get("v_min") and (v is None or v < crit["v_min"]): continue
        if crit.get("i_min") and (i is None or i < crit["i_min"]): continue
        if crit.get("mfr") and (r.get("mfr") or "") != crit["mfr"]: continue
        if crit.get("mounting") and (r.get("mounting") or "") != crit["mounting"]: continue
        if crit.get("package") and crit["package"] not in (r.get("package") or ""): continue
        if crit.get("tj_min") and (r.get("tj_max") is None or r.get("tj_max") < crit["tj_min"]): continue
        if crit.get("technology") and (r.get("technology") or "") != crit["technology"]: continue
        out.append(r)
    return out


# ── DB record → engine block (real scalars + labelled estimates) ──────────────
def _vf_curve(vf, vf_if):
    """A 2-point Vf(If) curve from the single datasheet (Vf @ If) point + a slope estimate."""
    if not vf or not vf_if: return None
    lo_i = max(0.5, vf_if * 0.2)
    lo_v = max(0.3, vf - 0.25)         # ~0.25 V drop from rated point to low current (estimate)
    return [[round(lo_i, 2), round(vf_if, 2)], [round(lo_v, 3), round(vf, 3)]]

_PKG_RTH = {"TO247": 0.5, "TO-247": 0.5, "TO264": 0.45, "TO220": 1.0, "TO-220": 1.0,
            "DPAK": 1.6, "D2PAK": 1.0, "DDPAK": 1.2, "TO252": 1.6, "TO-252": 1.6}
def _pkg_rth(rec, default):
    pk = (rec.get("supplier_package") or rec.get("package") or "").upper().replace(" ", "")
    for k, v in _PKG_RTH.items():
        if k.replace("-", "") in pk: return v
    return default

def to_block(rec, kind):
    """Map a DB record to an engine component block. Real datasheet scalars are used verbatim;
    missing loss/thermal parameters are estimated (marked in `_estimated`)."""
    est = []
    meta = {"manufacturer": rec.get("mfr"), "part_number": rec.get("part_number"),
            "datasheet_url": rec.get("datasheet_url")}
    if kind == "mosfet":
        sic = _is_sic(rec)
        rdson = rec.get("rdson") or 0.1
        qg = rec.get("qg") or 60e-9
        tjm = rec.get("tj_max") or 150
        # Rth_jc from rated power dissipation, else by package
        rth = (tjm - 25) / rec["pd_max"] if rec.get("pd_max") else _pkg_rth(rec, 0.9); est.append("rth_jc")
        # Eoss estimate: larger die (lower Rdson) ⇒ more Coss ⇒ more Eoss; ∝ 1/Rdson, ∝ V^1.5
        eoss400 = max(0.5e-6, 0.9e-6 / max(rdson, 1e-3)); est.append("eoss_at_v")
        blk = {**meta, "tech": "sic" if sic else "si",
               "rdson_25": rdson, "rdson_tj": [[25, 125], [1.0, 1.4]] if sic else [[25, 100], [1.0, 1.8]],
               "qg": qg, "qgd": qg * 0.25, "ciss": rec.get("ciss") or 1500e-12,
               "vth": rec.get("vth") or 4.0, "vpl": (rec.get("vth") or 4.0) + 2.0,
               "eoss_at_v": [[100, 400], [round(eoss400 * (0.25 ** 1.5), 9), round(eoss400, 9)]],
               "rth_jc": round(rth, 3), "rth_cs": 0.3, "vg": 15.0 if sic else 12.0, "rg": 4.0}
        est += ["rdson_tj", "qgd", "vpl"]
    elif kind == "diode":
        sic = _is_sic(rec)
        blk = {**meta, "is_sic": sic, "vf_curve": _vf_curve(rec.get("vf"), rec.get("vf_if")) or [[1, 5], [1.1, 1.7]],
               "vf_tco": 0.0015 if sic else -0.002,
               "rth_jc": _pkg_rth(rec, 0.8), "rth_cs": 0.3}
        if sic:
            blk["qc"] = 18e-9; est.append("qc")
        else:                                    # estimate Qrr ≈ ½·trr·Io
            trr = rec.get("trr") or 75e-9; io = rec.get("io") or 5
            blk["qrr"] = round(0.5 * trr * io, 12); est.append("qrr")
        est += ["rth_jc", "vf_curve(slope)"]
    else:  # bridge
        blk = {**meta, "topology": "diode",
               "vf_curve": _vf_curve(rec.get("vf"), rec.get("vf_if")) or [[1, 12], [0.8, 1.2]],
               "vf_tco": -0.002, "n_parallel": 1,
               "rth_jc": _pkg_rth(rec, 1.0), "rth_cs": 0.5}
        est += ["rth_jc", "vf_curve(slope)"]
    blk["_estimated"] = est
    return blk


# ── rank filtered candidates by computed loss at the design operating point ────
def rank_by_loss(kind, design, crit, top=10, max_eval=120):
    """Filter, then evaluate each candidate's loss across the 9 operating points and return the
    `top` lowest-loss parts. Returns [{part…, block, loss_W, tj_max_C}]."""
    from app.mode_b.semiconductor.adapter import build_semi_cfg
    from app.mode_b.semiconductor import pfc_loss_model as engine
    from app.mode_b.semiconductor.library import _SEED          # defaults for the other two blocks
    cands = filter_parts(kind, crit)
    # cheap pre-sort so we evaluate the most promising first: low Rdson / low Vf
    key = (lambda r: r.get("rdson") or 9) if kind == "mosfet" else (lambda r: r.get("vf") or 9)
    cands = sorted(cands, key=key)[:max_eval]
    loss_key = {"mosfet": "P_FET_total", "diode": "P_DIODE_total", "bridge": "P_BRIDGE_total"}[kind]
    tj_key = {"mosfet": "Tj_FET", "diode": "Tj_DIODE", "bridge": "Tj_BRIDGE_top"}[kind]
    base = {"mosfet": _SEED["mosfet"][0], "diode": _SEED["diode"][0], "bridge": _SEED["bridge"][0],
            "thermal": {"t_ambient": 45.0, "rth_sa": 0.35}}
    scored = []
    for rec in cands:
        blk = to_block(rec, kind)
        parts = {k: dict(base[k]) for k in ("mosfet", "diode", "bridge")}
        parts[kind] = blk
        try:
            cfg, _ = build_semi_cfg(design, parts["mosfet"], parts["diode"], parts["bridge"], base["thermal"])
            rows = [engine.simulate_point(float(v), *engine.design_from_dict(cfg)) for v in cfg["run"]["vac_list"]]
            loss = max(r[loss_key] for r in rows)
            tjm = max(r[tj_key] for r in rows)
        except Exception:
            continue
        scored.append({"manufacturer": rec.get("mfr"), "part_number": rec.get("part_number"),
                       "technology": rec.get("technology"), "package": rec.get("package"),
                       "mounting": rec.get("mounting"), "datasheet_url": rec.get("datasheet_url"),
                       "v_rating": rec.get("vdss") if kind == "mosfet" else rec.get("vr"),
                       "i_rating": rec.get("id_25") if kind == "mosfet" else rec.get("io"),
                       "loss_W": round(float(loss), 2), "tj_max_C": round(float(tjm), 1),
                       "block": blk})
    scored.sort(key=lambda x: x["loss_W"])
    return scored[:top]


if __name__ == "__main__":
    print("ingesting…", build_all())
