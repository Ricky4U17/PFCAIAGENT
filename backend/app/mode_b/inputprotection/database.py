"""
Local NTC inrush-current-limiter (ICL) database
===============================================
Parse the ICL_Database.xlsx parametric table, normalize each part, screen the catalog against the
NTC sizing result, and rank the survivors — the same "database agent" pattern the semiconductor
selector uses (`app.mode_b.semiconductor.database`).

The Excel (specs/Database/ICL_Database.xlsx) is a Digi-Key-style parametric table. It carries the
REAL selection scalars — R@25°C, tolerance, steady-state max current, hot resistance at rated
current, disc diameter, lead spacing, qualification (AEC-Q200), approval agency — but NOT the
pulse-energy rating (Joules) or the "max switchable capacitance" a vendor quotes for surge
survival. So the hard screen runs on the real R25 (drives the inrush limit); the pulse energy is
ESTIMATED from the disc diameter with a clearly-labelled correlation (marked in every reason and
in `_estimated`) and only ORDERS the survivors — the designer confirms energy / max-C on the
datasheet before ordering, exactly as the module docstring in ntc_bypass_select warns.

A LOCAL copy of the workbook is kept under ./data/ so the selector no longer depends on the
specs/ tree being present at runtime; `build_all()` refreshes both the copy and the JSON cache.
"""
from __future__ import annotations
import os, re, json, shutil

_HERE = os.path.dirname(os.path.abspath(__file__))
_DATA = os.path.join(_HERE, "data")
_SPEC = os.path.normpath(os.path.join(_HERE, "..", "..", "..", "..", "specs", "Database"))
# Surge-protection vendor workbooks (MOV + GDT) live under specs/Improvements/MOV.
_SURGE_SPEC = os.path.normpath(os.path.join(_HERE, "..", "..", "..", "..", "specs", "Improvements", "MOV"))
_XLSX = "ICL_Database.xlsx"
_JSON = "icl.json"

# Pulse-energy correlation (ESTIMATE): a high-energy disc's single-pulse capability scales with
# its bulk, ~ disc area. Calibrated against the representative large-disc parts (25 mm ~ 190 J,
# 30 mm ~ 270 J, 34 mm ~ 350 J → k ≈ 0.30 J/mm²). Conservative; always flagged as estimated.
_ENERGY_K = 0.30


# ── value helpers ─────────────────────────────────────────────────────────────
def _num(v):
    """Coerce a cell to float, tolerating '±20%', '5 A', stray text; None on failure."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    m = re.search(r"-?\d+\.?\d*", str(v))
    return float(m.group(0)) if m else None


def _energy_est_J(diameter_mm):
    """Estimated single-pulse energy capability from disc diameter (ESTIMATE, verify datasheet)."""
    d = _num(diameter_mm)
    return round(_ENERGY_K * d * d, 1) if d else None


# ── Excel ingest → normalized records ─────────────────────────────────────────
def _src_path():
    """Prefer the local copy under ./data; fall back to the specs/ workbook."""
    local = os.path.join(_DATA, _XLSX)
    return local if os.path.exists(local) else os.path.join(_SPEC, _XLSX)


def _rows(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h).strip() if h is not None else "" for h in next(it)]
    for r in it:
        if any(v not in (None, "") for v in r):
            yield dict(zip(hdr, r))
    wb.close()


def _pick(rec, *subs):
    """First value whose header contains all `subs` (case-insensitive) — robust to °C / unit noise."""
    for k, v in rec.items():
        kl = k.lower()
        if all(s.lower() in kl for s in subs):
            return v
    return None


def ingest():
    out = []
    for r in _rows(_src_path()):
        part = _pick(r, "part")
        if not part:
            continue
        d = _num(_pick(r, "diameter"))
        out.append({
            "mfr":            _pick(r, "mfr") if _pick(r, "mfr") else _pick(r, "manufacturer"),
            "part_number":    part,
            "series":         _pick(r, "series"),
            "description":    _pick(r, "description"),
            "r25":            _num(_pick(r, "r @ 25")) or _num(_pick(r, "25", "ohm")),
            "tolerance":      _pick(r, "tolerance"),
            "imax":           _num(_pick(r, "steady", "max")) or _num(_pick(r, "current", "max")),
            "r_hot_mohm":     _num(_pick(r, "r @ current")) or _num(_pick(r, "mohm")),
            "diameter_mm":    d,
            "lead_spacing_mm": _num(_pick(r, "lead", "spac")),
            "qualification":  _pick(r, "qualif"),
            "approval":       _pick(r, "approv"),
            "datasheet_url":  _pick(r, "datasheet"),
            "url":            _pick(r, "url"),
            "energy_est_J":   _energy_est_J(d),         # ESTIMATE from diameter
        })
    return out


def build_all():
    """Refresh the LOCAL workbook copy (from specs/) + the JSON cache. Returns the part count."""
    os.makedirs(_DATA, exist_ok=True)
    spec_xlsx = os.path.join(_SPEC, _XLSX)
    local_xlsx = os.path.join(_DATA, _XLSX)
    if os.path.exists(spec_xlsx) and os.path.abspath(spec_xlsx) != os.path.abspath(local_xlsx):
        shutil.copyfile(spec_xlsx, local_xlsx)          # keep a local copy of the excel
    recs = ingest()
    with open(os.path.join(_DATA, _JSON), "w", encoding="utf-8") as f:
        json.dump(recs, f)
    return len(recs)


# ── load + filter ─────────────────────────────────────────────────────────────
_CACHE = None
def load():
    global _CACHE
    if _CACHE is None:
        path = os.path.join(_DATA, _JSON)
        if not os.path.exists(path):                    # self-bootstrap on first use
            build_all()
        with open(path, encoding="utf-8") as f:
            _CACHE = json.load(f)
    return _CACHE


def options():
    """Distinct selectable values for GUI dropdowns."""
    recs = load()
    uniq = lambda key: sorted({(r.get(key) or "").strip() for r in recs if r.get(key)})
    return {"manufacturers": uniq("mfr"), "qualification": uniq("qualification"),
            "approval": uniq("approval")}


def filter_parts(crit=None):
    """crit: {r25_min, imax_min, mfr, qualification, diameter_max}."""
    crit = crit or {}
    out = []
    for r in load():
        if crit.get("r25_min") and (r.get("r25") is None or r["r25"] < crit["r25_min"]): continue
        if crit.get("imax_min") and (r.get("imax") is None or r["imax"] < crit["imax_min"]): continue
        if crit.get("diameter_max") and (r.get("diameter_mm") is None or r["diameter_mm"] > crit["diameter_max"]): continue
        if crit.get("mfr") and (r.get("mfr") or "") != crit["mfr"]: continue
        if crit.get("qualification") and (r.get("qualification") or "") != crit["qualification"]: continue
        out.append(r)
    return out


# ── screen the catalog against an NTC sizing result ───────────────────────────
def _label(rec):
    d = rec.get("diameter_mm"); r25 = rec.get("r25")
    bits = []
    if r25 is not None: bits.append(f"{r25:g}Ω")
    if d is not None:   bits.append(f"Ø{d:g}mm")
    tail = f" — {', '.join(bits)}" if bits else ""
    return f"{rec.get('mfr') or ''} {rec.get('part_number') or ''}".strip() + tail


def screen_catalog(s, r, top: int = 12):
    """Screen the vendor ICL catalog against the sizing result `r` (from ntc_bypass_select.compute).

    Returns rows (name, ok, reasons) — the SAME contract as the built-in representative catalog —
    with the qualifying parts ranked first (smallest adequate disc = most economical), then the
    closest near-misses. R25 is the REAL datasheet value; pulse energy is estimated from diameter.
    """
    scored = []
    for rec in load():
        reasons = []; ok = True
        r25 = rec.get("r25")
        if r25 is None:
            ok = False; reasons.append("no R25 on record")
        elif r25 < r.r25_required:
            ok = False; reasons.append(f"R25 {r25:g}Ω < {r.r25_required:.2f}Ω required (inrush too high)")
        e_est = rec.get("energy_est_J")
        if e_est is None:
            ok = False; reasons.append("no disc diameter → energy not estimable")
        elif e_est < r.e_pulse_required:
            ok = False
            reasons.append(f"energy ~{e_est:g} J (est. from Ø) < {r.e_pulse_required:.0f} J required — verify datasheet")
        else:
            reasons.append(f"energy ~{e_est:g} J (est. from Ø; confirm on datasheet)")
        imax = rec.get("imax")
        if ok and imax is not None and imax < r.i_rms_worst:
            reasons.append(f"note: I_max {imax:g}A < I_rms {r.i_rms_worst:.1f}A (OK — bypassed after precharge)")
        # rank key: passing first; then smallest adequate disc; then R25 nearest the pick
        d = rec.get("diameter_mm") or 1e6
        rank = (0 if ok else 1, d, abs((r25 or 1e6) - r.r25_pick))
        scored.append((rank, _label(rec), ok, reasons, rec))
    scored.sort(key=lambda x: x[0])
    return [(name, ok, reasons) for _, name, ok, reasons, _ in scored[:top]]


def find_part(part_number: str):
    """Exact part-number lookup (case-insensitive) in the ICL database. None if absent."""
    pn = (part_number or "").strip().lower()
    if not pn:
        return None
    for rec in load():
        if str(rec.get("part_number", "")).strip().lower() == pn:
            return rec
    return None


def selected_metrics(s, r, rec):
    """Recalculate the inrush design around a SPECIFIC selected NTC part.

    Uses the part's real R25 in place of the generic pick: actual cold inrush peak, precharge
    RC timing and bypass delay, plus the (estimated) pulse-energy margin. Returns a JSON-safe
    dict for the GUI card and the report's 'selected part' section."""
    r25 = rec.get("r25")
    if r25 is None:
        return None
    r_total   = float(r25) + r.r_parasitic
    i_inrush  = r.vin_pk_max / max(r_total, 1e-9)
    tau       = float(r25) * s.cout
    t_bypass  = s.tau_multiple * tau
    e_est     = rec.get("energy_est_J")
    e_margin  = (float(e_est) / r.e_cap) if (e_est and r.e_cap > 0) else None
    checks = {
        "r25_ok":    float(r25) >= r.r25_required,
        "energy_ok": (float(e_est) >= r.e_pulse_required) if e_est else None,
        "imax_note": (rec.get("imax") is not None and rec["imax"] < r.i_rms_worst),
    }
    return {
        "part_number":   rec.get("part_number"), "mfr": rec.get("mfr"),
        "r25_ohm":       float(r25),
        "imax_A":        rec.get("imax"),
        "diameter_mm":   rec.get("diameter_mm"),
        "energy_est_J":  e_est,
        "datasheet_url": rec.get("datasheet_url"),
        "i_inrush_actual_A": round(i_inrush, 2),
        "r_total_cold_ohm":  round(r_total, 3),
        "tau_ms":            round(tau * 1e3, 2),
        "t_bypass_ms":       round(t_bypass * 1e3, 1),
        "energy_margin":     round(e_margin, 2) if e_margin is not None else None,
        "checks":            checks,
        "meets_target":      i_inrush <= s.i_inrush_target * 1.001,
    }


def rank(s, r, top: int = 12):
    """Rich variant of screen_catalog: full records + verdict, for GUI cards / future use."""
    scored = []
    for rec in load():
        reasons = []; ok = True
        r25 = rec.get("r25"); e_est = rec.get("energy_est_J")
        if r25 is None or r25 < r.r25_required:
            ok = False; reasons.append("R25 below required")
        if e_est is None or e_est < r.e_pulse_required:
            ok = False; reasons.append("estimated energy below required")
        d = rec.get("diameter_mm") or 1e6
        scored.append(((0 if ok else 1, d), {**rec, "ok": ok, "reasons": reasons,
                       "energy_margin": (e_est / r.e_pulse_required) if e_est else None}))
    scored.sort(key=lambda x: x[0])
    return [rec for _, rec in scored[:top]]


# ══════════════════════════════════════════════════════════════════════════════
#  MOV (metal-oxide varistor) surge database
# ══════════════════════════════════════════════════════════════════════════════
# Same pattern as the ICL section, for the MOV surge selector (mov_surge_select). The vendor
# workbook columns match the shipped TEMPLATE (Part Number, Manufacturer, MCOV, V_1mA, Imax 8/20,
# Vc @ Imax, Energy 2ms, Disc Diameter, Type, Datasheet URL) — all REAL datasheet scalars, so the
# MOV screen needs NO estimates: MCOV, the V-I anchor (V_1mA), the 8/20 survival current and the
# clamp point all come straight from the record.
#
# The TEMPLATE is NOT treated as a live source — only a filled `MOV_Database.xlsx` / `mov_varistors
# .xlsx` counts — so until the designer drops a real file the richer built-in `MOV_CATALOG` stays
# the fallback (no regression). The moment a real file lands, `screen_catalog_mov` goes live.
# Live vendor sources (TEMPLATE excluded). The designer's curated combined file is preferred.
_MOV_XLSX = ["MOV_Combined_Database.xlsx", "MOV_Database.xlsx", "mov_varistors.xlsx"]
_MOV_JSON = "mov.json"


def _first_existing(names, *dirs):
    """First `dir/name` that exists, scanning dirs in order then names in order."""
    for d in dirs:
        for name in names:
            p = os.path.join(d, name)
            if os.path.exists(p):
                return p
    return None


def _mov_src_path():
    """A filled MOV workbook: local ./data copy first, else the surge-spec / legacy specs folders."""
    return _first_existing(_MOV_XLSX, _DATA, _SURGE_SPEC, _SPEC)


def _pkg_diameter_mm(pkg):
    """Disc diameter from a 'Disc 14mm' / 'Disc 10mm' package string (None if not a disc)."""
    if not pkg:
        return None
    m = re.search(r"(\d+\.?\d*)\s*mm", str(pkg))
    return float(m.group(1)) if m else None


def ingest_mov(path=None):
    """Normalize the MOV vendor workbook. The designer's MOV_Combined_Database.xlsx carries real
    datasheet scalars by EXACT header (MCOV, varistor voltage min/typ/max = V_1mA + tolerance, 8/20
    surge current, energy, capacitance, package); the max CLAMPING voltage (Vc @ In) is NOT exported,
    so `vc_imax` stays None → the clamp/let-through check is DATA-LIMITED (never a silent pass). The
    older _pick fallbacks keep the legacy template working."""
    src = path or _mov_src_path()
    if not src:
        return []
    out = []
    for r in _rows(src):
        part = r.get("Mfr Part #") or _pick(r, "part")
        if not part:
            continue
        v1ma = _num(r.get("Varistor Voltage Typ Numeric")) or _num(_pick(r, "v_1ma")) or _num(_pick(r, "1ma"))
        v1ma_min = _num(r.get("Varistor Voltage Min Numeric"))
        v1ma_max = _num(r.get("Varistor Voltage Max Numeric"))
        pkg = r.get("Package / Case") or _pick(r, "package")
        out.append({
            "mfr":           r.get("Mfr") or _pick(r, "manufacturer"),
            "part_number":   part,
            "series":        r.get("Series") or _pick(r, "series"),
            "description":   r.get("Description") or _pick(r, "description"),
            "mcov":          _num(r.get("Maximum AC Volts Numeric")) or _num(_pick(r, "mcov")),
            "vdc_max":       _num(r.get("Maximum DC Volts Numeric")),
            "v1ma":          v1ma,
            "v1ma_min":      v1ma_min,
            "v1ma_max":      v1ma_max,
            "imax":          _num(r.get("Surge Current A Numeric")) or _num(_pick(r, "imax")),
            "vc_imax":       _num(_pick(r, "vc", "imax")),   # absent in combined file → None (DATA MISSING)
            "energy_2ms_J":  _num(r.get("Energy J Numeric")) or _num(_pick(r, "energy")),
            "capacitance_pf": _num(r.get("Capacitance pF Numeric")),
            "op_temp":       r.get("Operating Temperature"),
            "package":       pkg,
            "diameter_mm":   _pkg_diameter_mm(pkg) or _num(_pick(r, "diameter")),
            "type":          r.get("Mounting Type") or _pick(r, "type"),
            "datasheet_url": r.get("Datasheet") or _pick(r, "datasheet"),
        })
    return out


def build_mov():
    """Refresh the local MOV workbook copy + JSON cache from a filled vendor file. 0 if none present."""
    src = _mov_src_path()
    if not src:
        return 0
    os.makedirs(_DATA, exist_ok=True)
    local = os.path.join(_DATA, os.path.basename(src))
    if os.path.abspath(src) != os.path.abspath(local):
        shutil.copyfile(src, local)
    recs = ingest_mov(local)
    with open(os.path.join(_DATA, _MOV_JSON), "w", encoding="utf-8") as f:
        json.dump(recs, f)
    return len(recs)


_MOV_CACHE = None
def load_mov():
    global _MOV_CACHE
    if _MOV_CACHE is None:
        path = os.path.join(_DATA, _MOV_JSON)
        if not os.path.exists(path):
            if not _mov_src_path():                     # no live file → let the engine fall back
                _MOV_CACHE = []
                return _MOV_CACHE
            build_mov()
        with open(path, encoding="utf-8") as f:
            _MOV_CACHE = json.load(f)
    return _MOV_CACHE


def options_mov():
    recs = load_mov()
    uniq = lambda key: sorted({(r.get(key) or "").strip() for r in recs if r.get(key)})
    return {"manufacturers": uniq("mfr"), "type": uniq("type")}


def _mov_label(rec):
    bits = []
    if rec.get("mcov") is not None: bits.append(f"{rec['mcov']:g}Vac")
    if rec.get("diameter_mm") is not None: bits.append(f"Ø{rec['diameter_mm']:g}mm")
    if rec.get("type"): bits.append(str(rec["type"]))
    tail = f" — {', '.join(bits)}" if bits else ""
    return f"{rec.get('mfr') or ''} {rec.get('part_number') or ''}".strip() + tail


def screen_catalog_mov(s, gov, mcov_req, pol, top: int = 12):
    """Screen the vendor MOV catalog for the GOVERNING path — same (name, ok, reasons) contract as
    mov_surge_select.screen_catalog, but data-driven. Returns [] when no live DB (engine falls back).

    All scalars are real datasheet values: MCOV vs the required class, 8/20 I_max with the 10-pulse
    repetitive derate vs I_sc, and the let-through solved on the part's OWN V-I curve (per-part alpha
    backed out of V_1mA / Vc@Imax) against the criterion device gate.
    """
    rows = screen_table_mov(s, gov, mcov_req, pol, top)
    return [(r["label"], r["ok"], r["reasons"]) for r in rows]


def _mcov_from_text(*texts):
    """Parse an 'NNN Vac' MCOV token from a description / part string (part-# consistency check)."""
    for t in texts:
        if not t:
            continue
        m = re.search(r"(\d{3,4})\s*v\s*ac", str(t).lower())
        if m:
            return float(m.group(1))
    return None


def screen_table_mov(s, gov, mcov_req, pol, top: int = 12):
    """Structured MOV candidate screen for the governing path — the expanded datasheet-column table the
    review asks for. Each row carries the real scalars (MCOV, V_1mA + tolerance, 8/20 I_max, energy,
    capacitance, package/size), the survival + clamp verdicts, a part-number-vs-MCOV consistency flag,
    and the overall verdict/reasons. Returns [] when no live DB (engine falls back to the builtin)."""
    recs = load_mov()
    if not recs:
        return []
    from . import mov_surge_select as mov
    v_drive = gov.v_oc + (mov.v_line_peak(s) if s.phase_superposition else 0.0)
    gate = s.device_absmax if pol.gate_uses_absmax else s.device_vds - pol.dev_margin_V
    BIG = 1e9
    scored = []
    for rec in recs:
        mcov = rec.get("mcov"); v1ma = rec.get("v1ma")
        imax = rec.get("imax"); vc_max = rec.get("vc_imax")
        row = {"label": _mov_label(rec), "part_number": rec.get("part_number"), "mfr": rec.get("mfr"),
               "mcov": mcov, "v1ma": v1ma, "v1ma_min": rec.get("v1ma_min"), "v1ma_max": rec.get("v1ma_max"),
               "imax": imax, "energy_2ms_J": rec.get("energy_2ms_J"),
               "capacitance_pf": rec.get("capacitance_pf"), "package": rec.get("package"),
               "diameter_mm": rec.get("diameter_mm"), "datasheet_url": rec.get("datasheet_url"),
               "clamp_vc": None, "clamp_status": "DATA MISSING", "part_num_consistent": None}
        if None in (mcov, v1ma, imax):
            row.update(ok=False, verdict="FAIL", reasons=["incomplete record (needs MCOV / V_1mA / I_max 8/20)"])
            scored.append(((3, BIG, 0.0), row)); continue
        reasons, ok, conditional = [], True, False   # ok=False only for a REAL violated limit; a missing
        #                                              datasheet field -> CONDITIONAL (still selectable).
        if mcov < mcov_req:
            ok = False; reasons.append(f"MCOV {mcov:g} < required {mcov_req:.0f} Vac")
        eff_imax = imax * s.repetitive_derate
        row["imax_derated"] = eff_imax
        if eff_imax < gov.i_sc:
            ok = False
            reasons.append(f"I_max {imax:g}A x{s.repetitive_derate:.2f} (10-pulse) = {eff_imax:.0f}A < I_sc {gov.i_sc:.0f}A")
        else:
            reasons.append(f"survival OK: I_max {imax:g}A x{s.repetitive_derate:.2f} = {eff_imax:.0f}A >= I_sc {gov.i_sc:.0f}A")
        # part-number consistency: any 'NNN Vac' token in the description/part must match the MCOV class.
        tok = _mcov_from_text(rec.get("description"), rec.get("part_number"))
        if tok is not None:
            row["part_num_consistent"] = abs(tok - mcov) <= 1
            if not row["part_num_consistent"]:
                ok = False
                reasons.append(f"part-# inconsistency: description implies {tok:g} Vac vs MCOV {mcov:g} Vac")
        # clamp / let-through — needs the datasheet Vc@In; absent in the combined export -> DATA MISSING.
        # A missing clamp does NOT fail the part (that would make EVERY part un-selectable) — it is
        # CONDITIONAL: selectable, but the ride-through cannot be confirmed until Vc@In is entered.
        if vc_max is None:
            clamp_rank = BIG
            conditional = True
            reasons.append("clamp/let-through: DATA MISSING (no Vc@In in DB) — CONDITIONAL: cannot confirm "
                           "downstream margin; add datasheet max-clamping voltage to verify ride-through")
        else:
            a_eff = mov.effective_alpha(v1ma, vc_max, imax)
            i_op, vc = mov.operating_point(v1ma, a_eff, v_drive, gov.z)
            clamp_rank = vc
            row["clamp_vc"] = vc; row["clamp_status"] = "computed"
            reasons.append(f"let-through ~{vc:.0f}V @ {i_op:.0f}A (drive {v_drive:.0f}V); gate {gate:.0f}V [crit {pol.name}]")
            if vc > gate:
                ok = False
                reasons.append(f"clamp {vc:.0f}V > gate {gate:.0f}V -> "
                               + ("cannot ride through (criterion A)" if pol.ride_through else "FAIL even for survival"))
            elif s.device_vds - pol.dev_margin_V < vc <= gate and not pol.ride_through:
                reasons.append(f"survives but bus disturbed -> unit resets (allowed under criterion {pol.name})")
        verdict = "FAIL" if not ok else ("CONDITIONAL" if conditional else "PASS")
        row.update(ok=ok, verdict=verdict, reasons=reasons)
        # rank: PASS first, then CONDITIONAL (data-limited), then FAIL; within tier best clamp / survival.
        pass_tier = 0 if verdict == "PASS" else (1 if verdict == "CONDITIONAL" else 2)
        scored.append(((pass_tier, clamp_rank, -eff_imax), row))
    scored.sort(key=lambda x: x[0])
    return [row for _, row in scored[:top]]


# ══════════════════════════════════════════════════════════════════════════════
#  GDT (gas-discharge tube) surge-diverter database
# ══════════════════════════════════════════════════════════════════════════════
# The designer's GDT_Combined_Database.xlsx carries, by EXACT header: DC sparkover nom/min/max, the
# ±tolerance (low/high %), the 8/20 impulse discharge current, pole count, a fail-short flag, package,
# and a datasheet URL. Two review-mandated fields are NOT in the export — the IMPULSE (dynamic)
# sparkover at a stated dv/dt, and the FOLLOW/HOLD current — so `v_impulse_spark` and `follow_current`
# stay None and the selector must treat them as DATA MISSING (a GDT on L-PE/N-PE without follow-current
# evidence is a FAIL, not a PASS — review §16/§17), never silently pass.
_GDT_XLSX = ["GDT_Combined_Database.xlsx", "GDT_Database.xlsx"]
_GDT_JSON = "gdt.json"


def _gdt_src_path():
    return _first_existing(_GDT_XLSX, _DATA, _SURGE_SPEC, _SPEC)


def ingest_gdt(path=None):
    src = path or _gdt_src_path()
    if not src:
        return []
    out = []
    for r in _rows(src):
        part = r.get("Mfr Part #") or _pick(r, "part")
        if not part:
            continue
        out.append({
            "mfr":            r.get("Mfr") or _pick(r, "manufacturer"),
            "part_number":    part,
            "description":    r.get("Description"),
            "v_spark_nom":    _num(r.get("DC Sparkover Nom V Numeric")) or _num(_pick(r, "sparkover", "nom")),
            "v_spark_min":    _num(r.get("DC Sparkover Min V Numeric")) or _num(_pick(r, "sparkover", "min")),
            "v_spark_max":    _num(r.get("DC Sparkover Max V Numeric")) or _num(_pick(r, "sparkover", "max")),
            "tolerance":      r.get("Tolerance"),
            "tol_low_pct":    _num(r.get("Tolerance Low % Numeric")),
            "tol_high_pct":   _num(r.get("Tolerance High % Numeric")),
            "imax_impulse":   _num(r.get("Impulse Discharge Current A Numeric")) or _num(_pick(r, "impulse", "numer")),
            "poles":          _num(r.get("Number of Poles")),
            "fail_short":     r.get("Fail Short"),
            "v_impulse_spark": None,        # NOT in export → DATA MISSING (impulse sparkover @ dv/dt)
            "follow_current":  None,        # NOT in export → DATA MISSING (follow/hold current)
            "package":        r.get("Package / Case") or _pick(r, "package"),
            "mounting":       r.get("Mounting Type"),
            "datasheet_url":  r.get("Datasheet") or _pick(r, "datasheet"),
        })
    return out


def build_gdt():
    """Refresh the local GDT workbook copy + JSON cache from the vendor file. 0 if none present."""
    src = _gdt_src_path()
    if not src:
        return 0
    os.makedirs(_DATA, exist_ok=True)
    local = os.path.join(_DATA, os.path.basename(src))
    if os.path.abspath(src) != os.path.abspath(local):
        shutil.copyfile(src, local)
    recs = ingest_gdt(local)
    with open(os.path.join(_DATA, _GDT_JSON), "w", encoding="utf-8") as f:
        json.dump(recs, f)
    return len(recs)


_GDT_CACHE = None
def load_gdt():
    global _GDT_CACHE
    if _GDT_CACHE is None:
        path = os.path.join(_DATA, _GDT_JSON)
        if not os.path.exists(path):
            if not _gdt_src_path():
                _GDT_CACHE = []
                return _GDT_CACHE
            build_gdt()
        with open(path, encoding="utf-8") as f:
            _GDT_CACHE = json.load(f)
    return _GDT_CACHE


def options_gdt():
    recs = load_gdt()
    uniq = lambda key: sorted({(r.get(key) or "").strip() for r in recs if r.get(key)})
    v = sorted({r["v_spark_nom"] for r in recs if r.get("v_spark_nom") is not None})
    return {"manufacturers": uniq("mfr"), "mounting": uniq("mounting"), "sparkover_nom_V": v}


def _gdt_label(rec):
    bits = []
    if rec.get("v_spark_nom") is not None: bits.append(f"{rec['v_spark_nom']:g}V DC-spark")
    if rec.get("imax_impulse") is not None: bits.append(f"{rec['imax_impulse']:g}A 8/20")
    if rec.get("poles") is not None: bits.append(f"{rec['poles']:g}-pole")
    tail = f" — {', '.join(bits)}" if bits else ""
    return f"{rec.get('mfr') or ''} {rec.get('part_number') or ''}".strip() + tail


def screen_table_gdt(gs, top: int = 12):
    """Structured GDT candidate screen for the common-mode paths, per the review's program logic:
    no-fire (min sparkover after tolerance vs swelled line peak), 8/20 surge-current class vs the design
    target, and dynamic (impulse) sparkover — the last is DATA MISSING in the combined export, so it is
    flagged, never assumed. Follow-current / fail-short are spec-level safety gates (see the engine).
    Returns [] when no live DB."""
    recs = load_gdt()
    if not recs:
        return []
    from . import gdt_surge_select as gdt
    _, _, i_req = gdt.resolve_stress(gs)
    i_req = i_req or 0.0
    BIG = 1e9
    scored = []
    for rec in recs:
        v_min = rec.get("v_spark_min")
        if v_min is None and rec.get("v_spark_nom") is not None and rec.get("tol_low_pct") is not None:
            v_min = rec["v_spark_nom"] * (1.0 + rec["tol_low_pct"] / 100.0)
        v_max = rec.get("v_spark_max")
        imax = rec.get("imax_impulse")
        nf = gdt.no_fire(gs, v_min)
        sc = gdt.surge_current(gs, i_req, imax)
        ds = gdt.dynamic_sparkover(gs, rec.get("v_impulse_spark"), v_max)
        reasons = [nf["note"], sc["note"], "dynamic sparkover: " + ds["note"]]
        ok = (nf["ok"] is True) and (sc["ok"] is True)     # ds unknown -> not a pass-blocker but flagged
        row = {"label": _gdt_label(rec), "part_number": rec.get("part_number"), "mfr": rec.get("mfr"),
               "v_spark_nom": rec.get("v_spark_nom"), "v_spark_min": v_min, "v_spark_max": v_max,
               "tolerance": rec.get("tolerance"), "imax_impulse": imax, "poles": rec.get("poles"),
               "fail_short": rec.get("fail_short"), "package": rec.get("package"),
               "impulse_spark": rec.get("v_impulse_spark"), "datasheet_url": rec.get("datasheet_url"),
               "no_fire_ok": nf["ok"], "surge_ok": sc["ok"], "dynamic_status": "DATA MISSING" if ds["ok"] is None else "ok",
               "ok": ok, "reasons": reasons}
        # rank: no-fire pass first, then smallest sufficient sparkover (fires earliest above no-fire), then highest current
        nf_tier = 0 if nf["ok"] else (2 if nf["ok"] is None else 1)
        scored.append(((nf_tier, 0 if sc["ok"] else 1, rec.get("v_spark_nom") or BIG, -(imax or 0)), row))
    scored.sort(key=lambda x: x[0])
    return [row for _, row in scored[:top]]


# ══════════════════════════════════════════════════════════════════════════════
#  Fuse (line-fuse) database
# ══════════════════════════════════════════════════════════════════════════════
# specs/Improvements/MOV/../FUSE/Fuse_Database.xlsx (115 parts). NOTE: the sheet has TITLE rows before
# the real header (first cell "Mfr Part #"), so this needs a header-skip reader — NOT the generic _rows()
# which assumes row 0 is the header. Real datasheet scalars: current rating, AC/DC voltage, AC/DC breaking
# capacity, melting I2t (25/115 blank -> DATA MISSING), response time, type, approvals.
_FUSE_XLSX = ["Fuse_Database.xlsx"]
_FUSE_JSON = "fuse.json"
_FUSE_SPEC = os.path.normpath(os.path.join(_HERE, "..", "..", "..", "..", "specs", "Improvements", "FUSE"))


def _fuse_src_path():
    return _first_existing(_FUSE_XLSX, _DATA, _FUSE_SPEC, _SURGE_SPEC, _SPEC)


def _fuse_rows(path):
    """Yield dict rows keyed by the real header (the first row whose first cell is 'Mfr Part #')."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Fuse Database"] if "Fuse Database" in wb.sheetnames else wb[wb.sheetnames[0]]
    hdr = None
    for row in ws.iter_rows(values_only=True):
        if hdr is None:
            if row and str(row[0]).strip() == "Mfr Part #":
                hdr = [str(h).strip() if h is not None else "" for h in row]
            continue
        if any(v not in (None, "") for v in row):
            yield dict(zip(hdr, row))
    wb.close()


def ingest_fuse(path=None):
    src = path or _fuse_src_path()
    if not src:
        return []
    out = []
    for r in _fuse_rows(src):
        part = r.get("Mfr Part #")
        if not part:
            continue
        out.append({
            "mfr":            r.get("Manufacturer"),
            "part_number":    part,
            "series":         r.get("Series"),
            "description":    r.get("Description"),
            "fuse_type":      r.get("Fuse Type"),
            "response_time":  r.get("Response Time"),
            "mounting":       r.get("Mounting Type"),
            "i_rated_A":      _num(r.get("Current Rating (A)")),
            "v_ac_V":         _num(r.get("Voltage AC (V)")),
            "v_dc_V":         _num(r.get("Voltage DC (V)")),
            "breaking_ac_A":  _num(r.get("Breaking Cap. AC (A)")),
            "breaking_dc_A":  _num(r.get("Breaking Cap. DC (A)")),
            "melting_i2t":    _num(r.get("Melting I²t (A²s)")),   # 25/115 blank -> None (DATA MISSING)
            "approval":       r.get("Approval Agency"),
            "op_temp":        r.get("Operating Temperature"),
            "package":        r.get("Package / Case"),
            "datasheet_url":  r.get("Datasheet"),
        })
    return out


def build_fuse():
    src = _fuse_src_path()
    if not src:
        return 0
    os.makedirs(_DATA, exist_ok=True)
    local = os.path.join(_DATA, os.path.basename(src))
    if os.path.abspath(src) != os.path.abspath(local):
        shutil.copyfile(src, local)
    recs = ingest_fuse(local)
    with open(os.path.join(_DATA, _FUSE_JSON), "w", encoding="utf-8") as f:
        json.dump(recs, f)
    return len(recs)


_FUSE_CACHE = None
def load_fuse():
    global _FUSE_CACHE
    if _FUSE_CACHE is None:
        path = os.path.join(_DATA, _FUSE_JSON)
        if not os.path.exists(path):
            if not _fuse_src_path():
                _FUSE_CACHE = []
                return _FUSE_CACHE
            build_fuse()
        with open(path, encoding="utf-8") as f:
            _FUSE_CACHE = json.load(f)
    return _FUSE_CACHE


def options_fuse():
    recs = load_fuse()
    uniq = lambda key: sorted({(r.get(key) or "").strip() for r in recs if r.get(key)})
    return {"manufacturers": uniq("mfr"), "fuse_type": uniq("fuse_type"), "response_time": uniq("response_time")}


def _fuse_label(rec):
    bits = []
    if rec.get("i_rated_A") is not None: bits.append(f"{rec['i_rated_A']:g}A")
    if rec.get("v_ac_V") is not None: bits.append(f"{rec['v_ac_V']:g}Vac")
    if rec.get("response_time"): bits.append(str(rec["response_time"]))
    tail = f" — {', '.join(bits)}" if bits else ""
    return f"{rec.get('mfr') or ''} {rec.get('part_number') or ''}".strip() + tail


def screen_table_fuse(fs, startup_i2t=None, top: int = 12):
    """Structured line-fuse screen against the vendor Fuse_Database. Criteria (fuse_select.requirements):
    AC voltage rating vs high line; current rating vs margin×I_rms (and not grossly oversized); breaking
    capacity vs available fault current (OPEN if none); melting I2t vs margin×startup I2t (no nuisance blow
    on the NTC-limited inrush) — melting I2t absent (25/115 parts) => DATA MISSING, never a silent pass."""
    recs = load_fuse()
    if not recs:
        return []
    from . import fuse_select as fz
    req = fz.requirements(fs, startup_i2t)
    BIG = 1e18
    scored = []
    _inr = req.get("inrush_peak")
    for rec in recs:
        i_rated = rec.get("i_rated_A"); v_ac = rec.get("v_ac_V")
        bc = rec.get("breaking_ac_A"); i2t = rec.get("melting_i2t")
        reasons, ok, conditional = [], True, False   # ok=False only for a REAL violated limit; a missing
        #                                              datasheet field -> CONDITIONAL (still selectable).
        # voltage
        v_ok = (v_ac is not None and v_ac >= req["v_min"])
        if not v_ok:
            ok = False
            reasons.append(f"V_ac {v_ac:g}V < {req['v_min']:.0f}V line" if v_ac is not None else "V_ac rating missing")
        # continuous + inrush current: I_rated must exceed BOTH the continuous margin AND the inrush peak
        i_ok = (i_rated is not None and i_rated >= req["i_rated_min"])
        if not i_ok:
            ok = False
            _why = f"cont {req['i_cont_min']:.1f}A" + (f" / inrush {_inr:.0f}A" if _inr else "")
            reasons.append(f"I_rated {i_rated:g}A < required {req['i_rated_min']:.1f}A ({_why})" if i_rated else "I_rated missing")
        elif req["i_rated_max"] and i_rated > req["i_rated_max"]:
            ok = False; reasons.append(f"I_rated {i_rated:g}A oversized (> {req['i_rated_max']:.0f}A) — won't clear a small overload")
        else:
            reasons.append(f"I_rated {i_rated:g}A OK (>= {req['i_rated_min']:.1f}A"
                           + (f", incl. inrush {_inr:.0f}A)" if _inr else ")"))
        # breaking capacity vs available fault current
        if req["bc_min"] is None:
            bc_ok = None; conditional = True; reasons.append("breaking-capacity check OPEN (available fault current not given)")
        elif bc is None:
            bc_ok = None; conditional = True; reasons.append("breaking capacity DATA MISSING")
        else:
            bc_ok = bc >= req["bc_min"]
            if not bc_ok:
                ok = False
            reasons.append(f"breaking {bc:g}A {'>=' if bc_ok else '<'} fault {req['bc_min']:.0f}A")
        # ride the NTC-limited startup inrush without nuisance blow (melting I2t)
        if i2t is None:
            i2t_ok = None; conditional = True   # DATA MISSING -> CONDITIONAL (selectable), not a hard FAIL
            reasons.append("melting I2t DATA MISSING — CONDITIONAL: cannot confirm no-nuisance-blow vs startup I2t")
        elif req["i2t_min"] is None:
            i2t_ok = None; reasons.append(f"melting I2t {i2t:g} A2s (startup I2t not given -> ride-inrush OPEN)")
        else:
            i2t_ok = i2t > req["i2t_min"]
            if not i2t_ok:
                ok = False
            reasons.append(f"melting I2t {i2t:g} A2s {'>' if i2t_ok else '<='} {req['i2t_min']:.1f} A2s (margin x startup)")
        verdict = "FAIL" if not ok else ("CONDITIONAL" if conditional else "PASS")
        row = {"label": _fuse_label(rec), "part_number": rec.get("part_number"), "mfr": rec.get("mfr"),
               "i_rated_A": i_rated, "v_ac_V": v_ac, "breaking_ac_A": bc, "melting_i2t": i2t,
               "response_time": rec.get("response_time"), "fuse_type": rec.get("fuse_type"),
               "package": rec.get("package"), "datasheet_url": rec.get("datasheet_url"),
               "v_ok": v_ok, "i_ok": i_ok, "bc_ok": bc_ok, "i2t_ok": i2t_ok, "ok": ok,
               "verdict": verdict, "reasons": reasons}
        # rank: PASS first, then CONDITIONAL, then FAIL; within tier smallest sufficient rating (tightest)
        pass_tier = 0 if verdict == "PASS" else (1 if verdict == "CONDITIONAL" else 2)
        scored.append(((pass_tier, i_rated or BIG, -(i2t or 0)), row))
    scored.sort(key=lambda x: x[0])
    return [row for _, row in scored[:top]]


if __name__ == "__main__":
    print("ingesting ICL database…", build_all(), "parts;  local copy + cache under", _DATA)
    print("ingesting MOV database…", build_mov(), "parts (0 = no filled vendor file yet; template ignored)")
    print("ingesting GDT database…", build_gdt(), "parts (0 = no filled vendor file yet)")
    print("ingesting Fuse database…", build_fuse(), "parts (0 = no filled vendor file yet)")
